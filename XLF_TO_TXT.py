import xml.etree.ElementTree as ET
import pandas as pd
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
import os
import re  # Import the regular expressions library
import polib  # Import the polib library for parsing .po files
from openpyxl import Workbook  # Import the openpyxl library for creating Excel files

def element_to_string(element):
    #Convert an XML element and its children to a string.
    if element is None:
        return ""
    parts = [element.text] + [ET.tostring(child, encoding='unicode') for child in element]
    # Filter out 'None' items, then concatenate the parts
    return ''.join(filter(None, parts)).strip()

# Function to remove XML tags from a string
def remove_xml_tags(text):
    # Define a regular expression pattern for XML tags and replace them with empty string
    pattern = re.compile(r'<.+?>')
    return pattern.sub('', text)

# Function to parse the .xlf file and extract source and target texts along with tags
def parse_xlf(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    # Namespaces handling
    namespaces = {
        'xliff': 'urn:oasis:names:tc:xliff:document:1.2',
        'html': 'http://www.w3.org/1999/xhtml'
    }
    
    source_lang = root.find('.//xliff:file', namespaces).attrib['source-language']
    target_lang = root.find('.//xliff:file', namespaces).attrib['target-language']
    
    source_texts = []
    target_texts = []
    
    for trans_unit in root.findall('.//xliff:trans-unit', namespaces):
        source = trans_unit.find('.//xliff:source', namespaces)
        target = trans_unit.find('.//xliff:target', namespaces)
        
        # Serialize the element, then remove tags from it
        source_text = remove_xml_tags(element_to_string(source))
        target_text = remove_xml_tags(element_to_string(target))
        
        source_texts.append(source_text)
        target_texts.append(target_text)
    
    return source_lang, target_lang, source_texts, target_texts

# Function to parse the .po file and extract source and target texts
def parse_po(file_path):
    po = polib.pofile(file_path)
    
    source_lang = po.metadata.get('Language')
    target_lang = 'en'  # Assuming the target language is English
    
    source_texts = []
    target_texts = []
    
    for entry in po:
        source_texts.append(entry.msgid)
        target_texts.append(entry.msgstr)
    
    return source_lang, target_lang, source_texts, target_texts

# Function to write source and target texts to a .txt file and an Excel file
def write_to_files(file_path, source_lang, target_lang, source_texts, target_texts):
    # Filter out any pairs where both source and target texts are empty or whitespace
    filtered_data = [(source, target) for source, target in zip(source_texts, target_texts)
                     if source.strip() or target.strip()]

    # Unzipping the filtered data to separate lists
    filtered_source_texts, filtered_target_texts = zip(*filtered_data) if filtered_data else ([], [])

    # Creating a DataFrame
    df = pd.DataFrame({
        source_lang: filtered_source_texts,
        target_lang: filtered_target_texts
    })
    
    # Write to .txt file
    txt_file_path = os.path.splitext(file_path)[0] + '.txt'
    df.to_csv(txt_file_path, sep='\t', index=False)
    
    # Write to Excel file
    excel_file_path = os.path.splitext(file_path)[0] + '.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Source and Target"
    
    ws.cell(row=1, column=1, value="Source")
    ws.cell(row=1, column=2, value="Target")
    
    for i, (source, target) in enumerate(zip(filtered_source_texts, filtered_target_texts), start=2):
        ws.cell(row=i, column=1, value=source)
        ws.cell(row=i, column=2, value=target)
    
    wb.save(excel_file_path)
    
    return txt_file_path, excel_file_path

# Function to prompt user for file selection
def choose_file():
    Tk().withdraw()  # Hide the root window
    file_path = askopenfilename(
        filetypes=[("Translation files", "*.xlf .sdlxliff .mxliff .mqxliff *.po"), ("All files", "*.*")]
    )
    return file_path

# Main function to handle the script workflow
def main():
    file_path = choose_file()
    if file_path:
        _, file_extension = os.path.splitext(file_path)
        if file_extension in ['.xlf', '.sdlxliff', '.mxliff', '.mqxliff']:
            source_lang, target_lang, source_texts, target_texts = parse_xlf(file_path)
        elif file_extension == '.po':
            source_lang, target_lang, source_texts, target_texts = parse_po(file_path)
        else:
            messagebox.showerror("Error", "Unsupported file format.")
            return
        
        txt_file_path, excel_file_path = write_to_files(file_path, source_lang, target_lang, source_texts, target_texts)
        messagebox.showinfo("Information", f"Files '{txt_file_path}' and '{excel_file_path}' created successfully!")

if __name__ == "__main__":
    main()